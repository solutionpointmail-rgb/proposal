"""
generate_excel.py
-----------------
Generates the Benefits Group proposal Excel workbook.
Matches the TennGreenLandConservancyProposalVertical.xlsx format.

Usage:
    python generate_excel.py
Output:
    Proposal_<ClientName>_<EffDate>.xlsx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime
from proposal_data import (
    CLIENT, PRODUCER, CONTRIBUTIONS,
    MEDICAL_PLANS, DENTAL_PLANS, VISION_PLANS, DISCLAIMER_TEXT
)

# ── COLOURS ───────────────────────────────────────────────────────────────────
TEAL       = "FF2E8B8B"   # header teal
LIGHT_TEAL = "FFE0F4F4"   # alternating row tint
DARK_TEXT  = "FF1A1A2E"
WHITE      = "FFFFFFFF"
ORANGE     = "FFFF6B35"   # accent / cost rows
LIGHT_GRAY = "FFF5F5F5"
MID_GRAY   = "FFD0D0D0"

def _font(bold=False, size=10, color="FF000000", name="Arial"):
    return Font(name=name, bold=bold, size=size, color=color)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border(style="thin"):
    s = Side(style=style, color="FFB0B0B0")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _money(val):
    return f"${val:,.2f}"

def _set(ws, row, col, value, bold=False, size=10, color="FF000000",
         fill=None, halign="left", wrap=True, border=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _font(bold=bold, size=size, color=color)
    cell.alignment = _align(h=halign, wrap=wrap)
    if fill:
        cell.fill = _fill(fill)
    if border:
        cell.border = _border()
    return cell

# ── INTRO SHEET ───────────────────────────────────────────────────────────────
def build_intro(wb):
    ws = wb.create_sheet("Intro")
    ws.sheet_properties.tabColor = "2E8B8B"
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 2
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 2
    ws.column_dimensions["F"].width = 20

    # Title block
    ws.row_dimensions[1].height = 6
    _set(ws, 2, 2, CLIENT["name"], bold=True, size=16, color=TEAL[2:])
    _set(ws, 3, 2, f"{CLIENT['address']}, {CLIENT['city_state_zip']}, County: {CLIENT['county']}", size=10)

    ws.row_dimensions[4].height = 6
    _set(ws, 5, 2, f"Prepared by: {PRODUCER['agency']} • {PRODUCER['sub_agency']}", bold=True, size=10)
    _set(ws, 5, 4, f"Total # of employees: {CLIENT['total_employees']}", bold=True)

    _set(ws, 6, 2, f"Effective date: {CLIENT['effective_date_mmddyyyy']}")
    _set(ws, 6, 4, f"Eligible Employees: {CLIENT['eligible_employees']}")
    _set(ws, 6, 6, f"SIC Code: {CLIENT['sic_code']}")

    _set(ws, 7, 4, f"Enrolling Employees: {CLIENT['enrolling_employees']}")
    _set(ws, 7, 6, f"Quote ID: {CLIENT['quote_id']}")

    ws.row_dimensions[8].height = 6
    _set(ws, 9, 2, "Employer Contribution", bold=True, size=11)
    _set(ws, 10, 2, f"Medical - {CONTRIBUTIONS['medical_ee']} / {CONTRIBUTIONS['medical_dep']}")
    _set(ws, 11, 2, f"Dental - {CONTRIBUTIONS['dental_ee']} / {CONTRIBUTIONS['dental_dep']}")
    _set(ws, 12, 2, f"Vision - {CONTRIBUTIONS['vision_ee']} / {CONTRIBUTIONS['vision_dep']}")
    _set(ws, 13, 2, f"Life - {CONTRIBUTIONS['life']}")

    ws.freeze_panes = None
    return ws

# ── MEDICAL SHEET ─────────────────────────────────────────────────────────────
def build_medical(wb):
    ws = wb.create_sheet("Medical")
    ws.sheet_properties.tabColor = "2E8B8B"

    plans = [p for p in MEDICAL_PLANS if p.get("include", True)]

    # Column layout: B=labels, then per plan: InNetwork + OutOfNetwork
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 32
    col_start = 3  # C
    for i, _ in enumerate(plans):
        in_col  = col_start + i * 2
        out_col = in_col + 1
        ws.column_dimensions[get_column_letter(in_col)].width  = 26
        ws.column_dimensions[get_column_letter(out_col)].width = 14

    # Section header
    _set(ws, 1, 2, "Medical Coverage", bold=True, size=14, color=TEAL[2:])

    # Plan headers
    r = 3
    for i, p in enumerate(plans):
        in_col = col_start + i * 2
        _set(ws, r,   in_col, str(i + 1), bold=True, size=11, halign="center",
             fill=TEAL, color=WHITE)
        _set(ws, r+1, in_col, p["plan_name"], bold=True, size=9,
             fill=LIGHT_TEAL, halign="center")
        _set(ws, r+2, in_col, f"{p['plan_type']} - {p['funding']}", size=9,
             fill=LIGHT_TEAL, halign="center")
        _set(ws, r+3, in_col, p["quote_type"], size=9,
             fill=LIGHT_TEAL, halign="center")

    # Column sub-headers
    r = 7
    _set(ws, r, 2, "Benefit Comparison", bold=True, fill=TEAL, color=WHITE)
    for i, _ in enumerate(plans):
        in_col  = col_start + i * 2
        out_col = in_col + 1
        _set(ws, r, in_col,  "In Network",     bold=True, fill=TEAL, color=WHITE, halign="center")
        _set(ws, r, out_col, "Out Of Network",  bold=True, fill=TEAL, color=WHITE, halign="center")

    # Benefit rows
    benefit_rows = [
        ("Deductible Individual",    "deductible_in_ind",  "deductible_out_ind"),
        ("Deductible Family",        "deductible_in_fam",  "deductible_out_fam"),
        ("Out-of-Pocket Max Individual", "oop_in_ind",     "oop_out_ind"),
        ("Out-of-Pocket Max Family", "oop_in_fam",         "oop_out_fam"),
        ("Coinsurance",              "coinsurance_in",      "coinsurance_out"),
        (None, None, None),  # spacer
        ("Doctor Visit",             "doctor_visit",        None),
        ("Speciality Visit",         "specialist",          None),
        ("Xray/Lab",                 "xray_lab",            None),
        ("Imaging",                  "imaging",             None),
        ("Urgent Care",              "urgent_care",         None),
        ("Emergency Room",           "er",                  None),
        ("Hospital Stay",            "hospital",            None),
        ("Prescription Drugs",       "rx",                  None),
        ("HSA Eligible",             "hsa_eligible",        None),
    ]

    r = 8
    for idx, (label, in_key, out_key) in enumerate(benefit_rows):
        bg = LIGHT_GRAY if idx % 2 == 0 else WHITE
        if label is None:
            r += 1
            continue
        _set(ws, r, 2, label, fill=bg, border=True)
        for i, p in enumerate(plans):
            in_col  = col_start + i * 2
            out_col = in_col + 1
            in_val = p.get(in_key, "")
            if isinstance(in_val, bool):
                in_val = "Yes" if in_val else "No"
            _set(ws, r, in_col, in_val, fill=bg, halign="center", border=True)
            if out_key:
                out_val = p.get(out_key, "")
                _set(ws, r, out_col, out_val, fill=bg, halign="center", border=True)
        r += 1

    # Cost section
    r += 1
    _set(ws, r, 2, "Cost Comparison", bold=True, fill=TEAL, color=WHITE)
    for i, p in enumerate(plans):
        in_col = col_start + i * 2
        _set(ws, r, in_col, p["quote_type"], bold=True, fill=TEAL, color=WHITE, halign="center")
    r += 1

    cost_rows = [
        (f"Employee Only ({CLIENT['tier_ee']})",   "rate_ee"),
        (f"Employee & Spouse ({CLIENT['tier_es']})", "rate_es"),
        (f"Employee & Child ({CLIENT['tier_ec']})",  "rate_ec"),
        (f"Employee & Family ({CLIENT['tier_ef']})", "rate_ef"),
    ]
    for idx, (label, key) in enumerate(cost_rows):
        bg = LIGHT_TEAL if idx % 2 == 0 else WHITE
        _set(ws, r, 2, label, fill=bg, border=True)
        for i, p in enumerate(plans):
            in_col = col_start + i * 2
            _set(ws, r, in_col, _money(p[key]), fill=bg, halign="right", border=True)
        r += 1

    r += 1
    for label, key, bg in [
        ("Employer Cost",    "employer_cost",  LIGHT_TEAL),
        ("Employee Cost",    "employee_cost",  WHITE),
        ("Monthly Premium",  "monthly_premium", ORANGE),
    ]:
        _set(ws, r, 2, label, bold=True, fill=bg, border=True)
        for i, p in enumerate(plans):
            in_col = col_start + i * 2
            color = WHITE if bg == ORANGE else "FF000000"
            _set(ws, r, in_col, _money(p[key]), bold=True,
                 fill=bg, halign="right", border=True, color=color)
        r += 1

    return ws

# ── DENTAL SHEET ──────────────────────────────────────────────────────────────
def build_dental(wb):
    ws = wb.create_sheet("Dental")
    ws.sheet_properties.tabColor = "FF6B35"
    plans = [p for p in DENTAL_PLANS if p.get("include", True)]

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 32
    col_start = 3
    for i, _ in enumerate(plans):
        ws.column_dimensions[get_column_letter(col_start + i * 2)].width     = 18
        ws.column_dimensions[get_column_letter(col_start + i * 2 + 1)].width = 14

    _set(ws, 1, 2, "Dental Coverage", bold=True, size=14, color="FFFF6B35")

    r = 3
    for i, p in enumerate(plans):
        c = col_start + i * 2
        _set(ws, r,   c, str(i + 1), bold=True, size=11, halign="center", fill=TEAL, color=WHITE)
        _set(ws, r+1, c, p["plan_name"], bold=True, size=9, fill=LIGHT_TEAL, halign="center")
        _set(ws, r+2, c, p["plan_type"], size=9, fill=LIGHT_TEAL, halign="center")

    r = 6
    _set(ws, r, 2, "Benefit Comparison", bold=True, fill=TEAL, color=WHITE)
    for i, _ in enumerate(plans):
        c = col_start + i * 2
        _set(ws, r, c,   "In Network",    bold=True, fill=TEAL, color=WHITE, halign="center")
        _set(ws, r, c+1, "Out Of Network", bold=True, fill=TEAL, color=WHITE, halign="center")

    dental_benefits = [
        ("Deductible Individual", "deductible_in_ind", "deductible_out"),
        ("Deductible Family",     "deductible_in_fam", "deductible_out"),
        ("Annual Maximum",        "annual_max_in",      "annual_max_out"),
    ]
    r = 7
    for idx, (label, in_key, out_key) in enumerate(dental_benefits):
        bg = LIGHT_GRAY if idx % 2 == 0 else WHITE
        _set(ws, r, 2, label, fill=bg, border=True)
        for i, p in enumerate(plans):
            c = col_start + i * 2
            _set(ws, r, c,   p.get(in_key, ""),  fill=bg, halign="center", border=True)
            _set(ws, r, c+1, p.get(out_key, ""), fill=bg, halign="center", border=True)
        r += 1

    r += 1
    _set(ws, r, 2, "Services Comparison", bold=True, fill=TEAL, color=WHITE)
    for i, _ in enumerate(plans):
        c = col_start + i * 2
        _set(ws, r, c,   "In Network",    bold=True, fill=TEAL, color=WHITE, halign="center")
        _set(ws, r, c+1, "Out Of Network", bold=True, fill=TEAL, color=WHITE, halign="center")
    r += 1

    services = [
        ("Preventive",          "preventive"),
        ("Basic",               "basic"),
        ("Major",               "major"),
        ("Ortho Maximum",       "ortho_max"),
        ("Endodontics",         "endodontics"),
        ("Simple Extractions",  "simple_extractions"),
        ("Implants",            "implants"),
        ("Periodontics",        "periodontics"),
        ("Oral Surgery",        "oral_surgery"),
        ("Composite Filling",   "composite_filling"),
        ("Waive Preventive Max","waive_preventive_max"),
    ]
    for idx, (label, key) in enumerate(services):
        bg = LIGHT_GRAY if idx % 2 == 0 else WHITE
        _set(ws, r, 2, label, fill=bg, border=True)
        for i, p in enumerate(plans):
            c = col_start + i * 2
            _set(ws, r, c, p.get(key, ""), fill=bg, halign="center", border=True)
            _set(ws, r, c+1, "See SBC",   fill=bg, halign="center", border=True)
        r += 1

    r += 1
    _set(ws, r, 2, "Cost Comparison", bold=True, fill=TEAL, color=WHITE)
    r += 1
    cost_rows = [
        (f"Employee Only ({CLIENT['tier_ee']})",    "rate_ee"),
        (f"Employee & Spouse ({CLIENT['tier_es']})", "rate_es"),
        (f"Employee & Child ({CLIENT['tier_ec']})",  "rate_ec"),
        (f"Employee & Family ({CLIENT['tier_ef']})", "rate_ef"),
    ]
    for idx, (label, key) in enumerate(cost_rows):
        bg = LIGHT_TEAL if idx % 2 == 0 else WHITE
        _set(ws, r, 2, label, fill=bg, border=True)
        for i, p in enumerate(plans):
            c = col_start + i * 2
            _set(ws, r, c, _money(p[key]), fill=bg, halign="right", border=True)
        r += 1

    r += 1
    for label, key, bg in [
        ("Employer Cost",   "employer_cost",   LIGHT_TEAL),
        ("Employee Cost",   "employee_cost",   WHITE),
        ("Monthly Premium", "monthly_premium", ORANGE),
    ]:
        _set(ws, r, 2, label, bold=True, fill=bg, border=True)
        for i, p in enumerate(plans):
            c = col_start + i * 2
            color = WHITE if bg == ORANGE else "FF000000"
            _set(ws, r, c, _money(p[key]), bold=True,
                 fill=bg, halign="right", border=True, color=color)
        r += 1

    return ws

# ── VISION SHEET ──────────────────────────────────────────────────────────────
def build_vision(wb):
    ws = wb.create_sheet("Vision")
    ws.sheet_properties.tabColor = "5B8DB8"
    plans = [p for p in VISION_PLANS if p.get("include", True)]

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 32
    col_start = 3
    for i, _ in enumerate(plans):
        ws.column_dimensions[get_column_letter(col_start + i)].width = 22

    _set(ws, 1, 2, "Vision Coverage", bold=True, size=14, color="FF5B8DB8")

    r = 3
    for i, p in enumerate(plans):
        c = col_start + i
        _set(ws, r,   c, str(i + 1), bold=True, size=11, halign="center", fill=TEAL, color=WHITE)
        _set(ws, r+1, c, p["plan_name"], bold=True, size=9, fill=LIGHT_TEAL, halign="center")
        _set(ws, r+2, c, p["plan_type"], size=9, fill=LIGHT_TEAL, halign="center")

    r = 6
    _set(ws, r, 2, "Benefit Frequency Comparison", bold=True, fill=TEAL, color=WHITE)
    r += 1

    freq_rows = [
        ("Exam",         "exam_frequency"),
        ("Lenses",       "lenses_frequency"),
        ("Frames",       "frames_frequency"),
        ("Contact Lens", "contacts_frequency"),
    ]
    for idx, (label, key) in enumerate(freq_rows):
        bg = LIGHT_GRAY if idx % 2 == 0 else WHITE
        _set(ws, r, 2, label, fill=bg, border=True)
        for i, p in enumerate(plans):
            _set(ws, r, col_start + i, p.get(key, "See SBC"), fill=bg, halign="center", border=True)
        r += 1

    r += 1
    _set(ws, r, 2, "Cost Comparison", bold=True, fill=TEAL, color=WHITE)
    r += 1

    cost_rows = [
        (f"Employee Only ({CLIENT['tier_ee']})",    "rate_ee"),
        (f"Employee & Spouse ({CLIENT['tier_es']})", "rate_es"),
        (f"Employee & Child ({CLIENT['tier_ec']})",  "rate_ec"),
        (f"Employee & Family ({CLIENT['tier_ef']})", "rate_ef"),
    ]
    for idx, (label, key) in enumerate(cost_rows):
        bg = LIGHT_TEAL if idx % 2 == 0 else WHITE
        _set(ws, r, 2, label, fill=bg, border=True)
        for i, p in enumerate(plans):
            _set(ws, r, col_start + i, _money(p[key]), fill=bg, halign="right", border=True)
        r += 1

    r += 1
    for label, key, bg in [
        ("Employer Cost",   "employer_cost",   LIGHT_TEAL),
        ("Employee Cost",   "employee_cost",   WHITE),
        ("Monthly Premium", "monthly_premium", ORANGE),
    ]:
        _set(ws, r, 2, label, bold=True, fill=bg, border=True)
        for i, p in enumerate(plans):
            color = WHITE if bg == ORANGE else "FF000000"
            _set(ws, r, col_start + i, _money(p[key]), bold=True,
                 fill=bg, halign="right", border=True, color=color)
        r += 1

    return ws

# ── DISCLAIMERS SHEET ─────────────────────────────────────────────────────────
def build_disclaimers(wb):
    ws = wb.create_sheet("Disclaimers")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 120
    _set(ws, 1, 2, "Disclaimers", bold=True, size=14, color=TEAL[2:])
    _set(ws, 3, 2, DISCLAIMER_TEXT, size=9, wrap=True)
    ws.row_dimensions[3].height = 120
    return ws

# ── MAIN ──────────────────────────────────────────────────────────────────────
def generate_excel(output_path=None):
    # Re-read from proposal_data at call time so app.py injections are picked up
    import proposal_data as _pd
    global CLIENT, CONTRIBUTIONS, MEDICAL_PLANS, DENTAL_PLANS, VISION_PLANS
    CLIENT        = _pd.CLIENT
    CONTRIBUTIONS = _pd.CONTRIBUTIONS
    MEDICAL_PLANS = _pd.MEDICAL_PLANS
    DENTAL_PLANS  = _pd.DENTAL_PLANS
    VISION_PLANS  = _pd.VISION_PLANS
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    build_intro(wb)
    build_medical(wb)
    build_dental(wb)
    build_vision(wb)
    build_disclaimers(wb)

    if output_path is None:
        safe_name = CLIENT["name"].replace(" ", "_").replace("/", "-")
        eff       = CLIENT["effective_date_mmddyyyy"].replace("/", "")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"/mnt/user-data/outputs/Proposal_{safe_name}_{eff}_v{timestamp}.xlsx"

    wb.save(output_path)
    print(f"✅ Excel saved → {output_path}")
    return output_path

if __name__ == "__main__":
    generate_excel()
